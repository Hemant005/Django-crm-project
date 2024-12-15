from django.shortcuts import render,redirect,get_object_or_404
from .forms import Product_Form,Client_Form,Opportunity_Form,Invoice_Form,Shipping_Form
from .models import Product,Client,Opportunity,Invoice,Shipping_Receipt,UserProfile
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.contrib.auth.models import User
from django.http import HttpResponse
from openpyxl import Workbook
from io import BytesIO
from fpdf import FPDF
import base64
import matplotlib
matplotlib.use('Agg') 
import matplotlib.pyplot as plt
import textwrap

# Create your views here.
def home(request):
    return render(request ,'crm/home.html')


@login_required
def dashboard_view(request):
    products_count=len(Product.objects.all())
    customer_count=len(Client.objects.all())
    shipping_count=len(Shipping_Receipt.objects.all())
    Opportunity_count=len(Opportunity.objects.all())
    
    top_products = Product.objects.all().order_by('-product_qty')[:5]
    product_names = [product.product_name for product in top_products]
    product_quantities = [product.product_qty for product in top_products]

    wrapped_product_names = [textwrap.fill(name, 10) for name in product_names] 

        # Create the bar chart using Matplotlib
    fig, ax = plt.subplots()
    ax.bar(product_names, product_quantities, color='blue',width=0.5)
    ax.set_xlabel('Product Name')
    ax.set_ylabel('Quantity')
    ax.set_title('Top 5 Products by Quantity')

    ax.set_xticks(range(len(product_names)))

    ax.set_xticklabels(wrapped_product_names)

    fig.tight_layout()

        # Save the plot to a BytesIO object
    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    image_png = buffer.getvalue()
    buffer.close()

        # Encode the image to base64 and pass it to the template
    graph = base64.b64encode(image_png).decode('utf-8')

    context={
        'products_count':products_count,
        'customer_count':customer_count,
        'shipping_count':shipping_count,
        'Opportunity_count':Opportunity_count,
        'graph':graph,
    }

    return render(request,'crm/dashboard.html',context)

def Logout_View(request):
    logout(request)
    return render(request, 'crm/logout.html')

@login_required
def add_client(request):
    user = request.user
    
    # Only allow Sales and Sales Managers groups to add clients
    if user.groups.filter(name='Sales').exists() or user.groups.filter(name='Sales Managers').exists():
        if request.method == 'POST':
            form = Client_Form(request.POST)
            if form.is_valid():
                client = form.save(commit=False)
                client.added_by = user  # Set the user who added the client
                client.save()
                return redirect('clients')  # Redirect to the clients page
        else:
            form = Client_Form()
        return render(request, 'crm/client_form.html', {'form': form})
    else:
        return render(request, 'crm/unauthorised.html')  # Show an unauthorized access page

@login_required
def add_product(request):
    user=request.user
    
    if user.groups.filter(name='Inventory').exists():
        if request.method == 'POST':
         form = Product_Form(request.POST)
         if form.is_valid():
             product=form.save(commit=False)
             product.added_by = user
             product.save()
             return redirect('products')
        else:
            form=Product_Form()
        return render(request,'crm/product_form.html',{'form':form})
    else:
        return render(request,'crm/unauthorised.html')

    
@login_required
def add_opportunity(request):
    user=request.user
    if user.groups.filter(name='Sales').exists():
        if request.method == 'POST':
            form = Opportunity_Form(request.POST)
            if form.is_valid():
                form.instance.added_by = user
                form.save()
                return redirect('opportunities')
        else:
            form = Opportunity_Form()
        return render(request, 'crm/opportunity_form.html', {'form': form})
    else:
        # Redirect or show an error if the user doesn't have permission
        return render(request,'crm/unauthorised.html')  # Change to your error page or dashboard

@login_required
def create_invoice(request):
    user = request.user
    
    # Check if the user belongs to the 'Finance' group (adjust the group name as needed)
    if user.groups.filter(name='Accounts').exists():
        if request.method == 'POST':
            form = Invoice_Form(request.POST)
            if form.is_valid():
                form.instance.added_by = user
                form.save()
                return redirect('invoices')
        else:
            form = Invoice_Form()
        
        return render(request, 'crm/invoice_form.html', {'form': form})
    else:
        return render(request, 'crm/unauthorised.html')

@login_required
def create_shipping(request):
    user=request.user
    if user.groups.filter(name='Inventory').exists():
        if request.method == 'POST':
            form = Shipping_Form(request.POST)
            if form.is_valid():
                form.instance.added_by=user
                form.save()
                return redirect('shippings')
        else:
            form=Shipping_Form()
        return render(request,'crm/shipping_form.html',{'form':form})
    else:
        return render(request,'crm/unauthorised.html')

    # if request.method == 'POST':
    #     form = Shipping_Form(request.POST)
    #     if form.is_valid():
    #         form.instance.added_by=request.user
    #         form.save()
    #         return redirect('shippings')
    # else:
    #     form=Shipping_Form()
    # return render(request,'crm/shipping_form.html',{'form':form})

@login_required
def clients_view(request):
    user = request.user

    # Check if the user is part of the Sales Managers group
    if user.groups.filter(name='Sales Manager').exists():
        # Managers can see all clients
        clients = Client.objects.all()
    elif user.groups.filter(name='Sales').exists():
        # Regular sales employees can only see their own clients
        clients = Client.objects.filter(added_by=user)
    elif user.groups.filter(name='Accounts').exists() or user.groups.filter(name='Inventory').exists():
        # Accounts and Inventory groups can see all clients but cannot add them
        clients = Client.objects.all()
    else:
        # Handle users not in any of the specified groups
        clients = Client.objects.none()

    client_count = clients.count()
    return render(request, "crm/clients.html", {'Clients': clients, 'Client_Count': client_count})

@login_required
def products_view(request):
    Products=Product.objects.all()
    Product_count=len(Product.objects.all())
    return render(request,"crm/products.html",{'Products':Products,'product_Count':Product_count})

@login_required
def opportunities_view(request):
    user=request.user

    if user.groups.filter(name='Sales Manager').exists() or user.groups.filter(name='Accounts').exists():
        opportunities=Opportunity.objects.all()
    elif user.groups.filter(name='Sales').exists():
        opportunities=Opportunity.objects.filter(added_by=user)
    else:
        opportunities=Opportunity.objects.none()
    
    opportunity_count=len(Opportunity.objects.all())
    return render(request,"crm/opportunities.html",{'Opportunities':opportunities,'opportunity_Count':opportunity_count})

@login_required
def invoice_view(request):
    user = request.user
    # Check if the user is in the Accounts group or the Sales Manager group
    if user.groups.filter(name='Accounts').exists() or user.groups.filter(name='Sales Manager').exists() or user.groups.filter(name='Inventory').exists():
        invoices = Invoice.objects.all()  # Can see all invoices
    elif user.groups.filter(name='Sales').exists():
        invoices = Invoice.objects.filter(opportunity__added_by=user)  # Only see invoices for their opportunities
    else:
        invoices = Invoice.objects.none()  # No access

    invoice_count = invoices.count()  # More efficient counting
    return render(request, "crm/invoices.html", {'Invoices': invoices, 'Invoice_Count': invoice_count})

@login_required
def Shipping_view(request):
    user = request.user

    # Check if the user is part of the Sales Manager group
    if user.groups.filter(name='Sales Manager').exists() or user.groups.filter(name='Accounts').exists():
        Shippings = Shipping_Receipt.objects.all()  # Sales Managers and accounts can see all shipments
    # Check if the user is part of the Inventory group
    elif user.groups.filter(name='Inventory').exists():
        Shippings = Shipping_Receipt.objects.all()  # Inventory can also see all shipments
    elif user.groups.filter(name='Sales').exists():
        Shippings = Shipping_Receipt.objects.filter(opportunity__added_by=user)  # Other users see only their shipments
    else:
        Shippings=Shipping_Receipt.objects.none()
    
    Shipping_count = Shippings.count()  # Get the count based on the filtered queryset
    return render(request, "crm/shippings.html", {'Shippings': Shippings, 'Shipping_count': Shipping_count})


#Edit and delete views
@login_required
def edit_client(request, client_id):
    user = request.user
    client = Client.objects.get(id=client_id)

    # Check if the user is the one who added the client
    if user == client.added_by:
        if request.method == 'POST':
            form = Client_Form(request.POST, instance=client)
            if form.is_valid():
                form.save()
                return redirect('clients')  # Redirect to the clients page after editing
        else:
            form = Client_Form(instance=client)
        return render(request, 'crm/client_form.html', {'form': form})
    else:
        return render(request, 'crm/unauthorised.html')  # Unauthorized access page

@login_required
def delete_client(request, client_id):
    user = request.user
    client = Client.objects.get(id=client_id)

    # Check if the user is the one who added the client
    if user == client.added_by:
        if request.method == 'POST':
            client.delete()
            return redirect('clients')  # Redirect to the clients page after deletion
        return render(request, 'crm/delete_client_confirmation.html', {'client': client})
    else:
        return render(request, 'crm/unauthorised.html')  # Unauthorized access page

@login_required
def edit_opportunity(request, opportunity_id):
    user = request.user
    opportunity = Opportunity.objects.get(id=opportunity_id)

    # Check if the user is the one who added the opportunity
    if user == opportunity.added_by:
        if request.method == 'POST':
            form = Opportunity_Form(request.POST, instance=opportunity)
            if form.is_valid():
                form.save()
                return redirect('opportunities')  # Redirect to the opportunities page after editing
        else:
            form = Opportunity_Form(instance=opportunity)
        return render(request, 'crm/opportunity_form.html', {'form': form})
    else:
        return render(request, 'crm/unauthorised.html')  # Unauthorized access page


@login_required
def delete_opportunity(request, opportunity_id):
    user = request.user
    opportunity = Opportunity.objects.get(id=opportunity_id)

    # Check if the user is the one who added the opportunity
    if user == opportunity.added_by:
        if request.method == 'POST':
            opportunity.delete()
            return redirect('opportunities')  # Redirect to the opportunities page after deletion
        return render(request, 'crm/delete_opportunity_confirmation.html', {'opportunity': opportunity})
    else:
        return render(request, 'crm/unauthorised.html')  # Unauthorized access page
    
@login_required
def edit_invoice(request, invoice_id):
    user = request.user
    invoice = Invoice.objects.get(id=invoice_id)

    # Check if the user is the one who added the invoice
    if user == invoice.added_by:
        if request.method == 'POST':
            form = Invoice_Form(request.POST, instance=invoice)
            if form.is_valid():
                form.save()
                return redirect('invoices')  # Redirect to the invoices page after editing
        else:
            form = Invoice_Form(instance=invoice)
        return render(request, 'crm/invoice_form.html', {'form': form})
    else:
        return render(request, 'crm/unauthorised.html')  # Unauthorized access page

@login_required
def delete_invoice(request, invoice_id):
    user = request.user
    invoice = Invoice.objects.get(id=invoice_id)

    # Check if the user is the one who added the invoice
    if user == invoice.added_by:
        if request.method == 'POST':
            invoice.delete()
            return redirect('invoices')  # Redirect to the invoices page after deletion
        return render(request, 'crm/delete_invoice_confirmation.html', {'invoice': invoice})
    else:
        return render(request, 'crm/unauthorised.html')  # Unauthorized access page

@login_required
def edit_product(request, product_id):
    user = request.user
    product = Product.objects.get(id=product_id)

    # Check if the user is the one who added the product
    if user == product.added_by:
        if request.method == 'POST':
            form = Product_Form(request.POST, instance=product)
            if form.is_valid():
                form.save()
                return redirect('products')  # Redirect to the products page after editing
        else:
            form = Product_Form(instance=product)
        return render(request, 'crm/product_form.html', {'form': form})
    else:
        return render(request, 'crm/unauthorised.html')  # Unauthorized access page

@login_required
def delete_product(request, product_id):
    user = request.user
    product = Product.objects.get(id=product_id)

    # Check if the user is the one who added the product
    if user == product.added_by:
        if request.method == 'POST':
            product.delete()
            return redirect('products')  # Redirect to the products page after deletion
        return render(request, 'crm/delete_product_confirmation.html', {'product': product})
    else:
        return render(request, 'crm/unauthorised.html')  # Unauthorized access page

@login_required
def edit_shipping(request, shipping_id):
    user = request.user
    shipping = Shipping_Receipt.objects.get(id=shipping_id)

    # Check if the user is the one who added the shipping receipt
    if user == shipping.added_by:
        if request.method == 'POST':
            form = Shipping_Form(request.POST, instance=shipping)
            if form.is_valid():
                form.save()
                return redirect('shippings')  # Redirect to the shippings page after editing
        else:
            form = Shipping_Form(instance=shipping)
        return render(request, 'crm/shipping_form.html', {'form': form})
    else:
        return render(request, 'crm/unauthorised.html')  # Unauthorized access page

@login_required
def delete_shipping(request, shipping_id):
    user = request.user
    shipping = Shipping_Receipt.objects.get(id=shipping_id)

    # Check if the user is the one who added the shipping receipt
    if user == shipping.added_by:
        if request.method == 'POST':
            shipping.delete()
            return redirect('shippings')  # Redirect to the shippings page after deletion
        return render(request, 'crm/delete_shipping_confirmation.html', {'shipping': shipping})
    else:
        return render(request, 'crm/unauthorised.html')  # Unauthorized access page


@login_required
def profile_view(request):
    # Get the user profile for the logged-in user
    profile = get_object_or_404(UserProfile, user=request.user)
    
    return render(request, 'crm/profile_view.html', {'profile': profile})

@login_required
def export_invoices(request):
    # Check if the user belongs to the 'Accounts' group
    if request.user.groups.filter(name='Accounts').exists():  
        # Create a workbook and a worksheet
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Invoices'

        # Define the header
        headers = [
            'Invoice ID', 'Client', 'Opportunity', 'Date', 'Product', 'Quantity', 
            'Total Amount (₹)', 'Payment Date', 'Payment Method', 'Status', 
            'Product Tax', 'Total with Tax', 'Added By'
        ]
        worksheet.append(headers)

        # Fetch the invoice data
        invoices = Invoice.objects.all()
        for invoice in invoices:
            worksheet.append([
                invoice.id,
                invoice.client.name,
                invoice.opportunity.opportunity_name,
                invoice.date,
                invoice.product.product_name,
                invoice.quantity,
                invoice.total_amount,
                invoice.payment_date if invoice.payment_date else "N/A",
                invoice.payment_method if invoice.payment_method else "N/A",
                invoice.status,
                invoice.product_tax(),
                invoice.total_with_tax(),
                invoice.added_by.username
            ])
        # Save the workbook to a BytesIO object
        output = BytesIO()
        workbook.save(output)
        output.seek(0)  # Move the cursor to the beginning of the stream


        # Create an HTTP response with the Excel file
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="invoices.xlsx"'
        
        return response
    else:
        return render(request, 'crm/unauthorised.html')
    
#View to generate Invoice PDF
@login_required
def export_invoice_pdf(request, invoice_id):
    # Fetch the specific invoice using the provided ID
    invoice = get_object_or_404(Invoice, id=invoice_id)

    # Create a PDF document
    pdf = FPDF()
    pdf.add_page()

    # Set font for the PDF
    pdf.set_font("Arial", size=12)

    # Add invoice details
    pdf.cell(200, 10, txt=f"ABC Corp", align='C')
    pdf.cell(200, 10, txt=f"Invoice Details")
    pdf.cell(200, 10, txt=f"Opportunity: {invoice.opportunity.opportunity_name}", ln=True)
    pdf.cell(200, 10, txt=f"Date: {invoice.date}", ln=True)
    pdf.cell(200, 10, txt=f"Client: {invoice.client.name}", ln=True)
    pdf.cell(200, 10, txt=f"Total Amount: Rs.{invoice.total_amount}", ln=True)
    pdf.cell(200, 10, txt=f"Status: {invoice.status}", ln=True)

    # Add product details (assuming there's a related product field)
    pdf.cell(200, 10, txt="Products:", ln=True)
    pdf.set_font("Arial", size=10)
    pdf.cell(50, 10, txt="Product", border=1,align='C')
    pdf.cell(28, 10, txt="Quantity", border=1,align='C')
    pdf.cell(28, 10, txt="Price", border=1,align='C')
    pdf.cell(28, 10, txt="Total Price", border=1,align='C')
    pdf.cell(28, 10, txt="Tax", border=1,align='C')
    pdf.cell(28, 10, txt="Total with Tax", border=1,align='C')
    pdf.ln()

    # Assuming invoice has a related name 'products' that returns a list of products
    product = invoice.product
    pdf.cell(50, 10, txt=product.product_name, border=1)
    pdf.cell(28, 10, txt=str(invoice.quantity), border=1)  # Use quantity from the invoice
    pdf.cell(28, 10, txt=f"Rs.{product.cost_price}", border=1)  # Use cost_price of the product
    pdf.cell(28, 10, txt=f"Rs.{invoice.total_amount}", border=1)
    pdf.cell(28, 10, txt=f"Rs.{invoice.product_tax()}", border=1)  # Use the tax calculated from the invoice
    pdf.cell(28, 10, txt=f"Rs.{invoice.total_with_tax()}", border=1)
    pdf.ln()
    # for product in invoice.product.all():
    #     pdf.cell(50, 10, txt=product.product_name, border=1)
    #     pdf.cell(30, 10, txt=str(product.quantity), border=1)
    #     pdf.cell(30, 10, txt=f"₹{product.price}", border=1)
    #     pdf.cell(30, 10, txt=f"₹{product.tax}", border=1)
    #     pdf.ln()

    # Create the response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="invoice_{invoice.id}.pdf"'
    
    # Output the PDF to the response
    pdf_output = pdf.output(dest='S').encode('latin1')  # Encode to handle special characters
    response.write(pdf_output)


    return response


@login_required
def top_orders_view(request):
    user = request.user

    # Check if the user is a sales manager
    if user.groups.filter(name='Sales Manager').exists() or user.groups.filter(name='Accounts').exists():
        # Get the top 5 invoices by total_amount
        top_invoices = Invoice.objects.all().order_by('-total_amount')[:5]
        client_names = [invoice.client.name for invoice in top_invoices]
        total_amounts = [invoice.total_amount for invoice in top_invoices]

        # Wrap long client names to fit within the plot
        wrapped_client_names = [textwrap.fill(name, 10) for name in client_names]

        # Create the bar chart using Matplotlib
        fig, ax = plt.subplots(figsize=(10, 6))  # Adjust figure size if necessary
        ax.bar(wrapped_client_names, total_amounts, color='green', width=0.5)  # Reduced bar width

        # Set labels and title
        ax.set_xlabel('Client Name')
        ax.set_ylabel('Total Invoice Amount')
        ax.set_title('Top 5 Orders by Invoice Amount')

        # Adjust layout and reduce whitespace
        fig.tight_layout()

        # Save the plot to a BytesIO object
        buffer = BytesIO()
        fig.savefig(buffer, format='png')
        buffer.seek(0)
        image_png = buffer.getvalue()
        buffer.close()

        # Encode the image to base64 and pass it to the template
        graph = base64.b64encode(image_png).decode('utf-8')

        context = {
            'graph': graph  # Add the graph to the context
        }

        return render(request, 'crm/top_orders.html', context)

    else:
        return render(request, 'crm/unauthorised.html')
    

